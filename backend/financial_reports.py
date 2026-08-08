import calendar
import io
import json
import os
import posixpath
import re
import shutil
import tempfile
from datetime import datetime
from io import BytesIO
from typing import Optional
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload
from starlette.background import BackgroundTask

import backend.auth as auth
import backend.models as models
import backend.schemas as schemas
from backend.database import SQLALCHEMY_DATABASE_URL, get_db
from backend.time_utils import build_ph_date_range_bounds, get_ph_today


router = APIRouter(tags=["Financial Reports"])

FINANCIAL_REPORT_ROLES = {"admin", "administrator", "staff"}
DEFAULT_TEMPLATE_FILENAME = "CANTEEN-REPORT-2025-2026-2 (1).xlsx"
TEMPLATE_FILENAME = DEFAULT_TEMPLATE_FILENAME
TEMPLATE_DIR = os.path.join(os.path.abspath(os.path.dirname(__file__)), "report_templates")
TEMPLATES_CONFIG_PATH = os.path.join(TEMPLATE_DIR, "templates_config.json")


def _get_templates_config() -> dict:
    try:
        os.makedirs(TEMPLATE_DIR, exist_ok=True)
        if os.path.exists(TEMPLATES_CONFIG_PATH):
            with open(TEMPLATES_CONFIG_PATH, "r", encoding="utf-8") as f:
                config = json.load(f)
                if isinstance(config, dict) and "templates" in config and isinstance(config["templates"], list):
                    if "active_filename" not in config or not config["active_filename"]:
                        config["active_filename"] = DEFAULT_TEMPLATE_FILENAME
                    return config
    except Exception as exc:
        print(f"Error reading templates config: {exc}")

    initial = {
        "active_filename": DEFAULT_TEMPLATE_FILENAME,
        "templates": [
            {
                "filename": DEFAULT_TEMPLATE_FILENAME,
                "name": "Default DepEd Canteen Report Template",
                "is_default": True,
                "uploaded_at": "2026-01-01T00:00:00Z",
            }
        ],
    }
    try:
        _save_templates_config(initial)
    except Exception:
        pass
    return initial


def _save_templates_config(config: dict) -> None:
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    with open(TEMPLATES_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)


MONTH_SEQUENCE = [
    (0, 6, "June"),
    (1, 7, "July"),
    (2, 8, "August"),
    (3, 9, "September"),
    (4, 10, "October"),
    (5, 11, "November"),
    (6, 12, "December"),
    (7, 1, "January"),
    (8, 2, "February"),
    (9, 3, "March"),
    (10, 4, "April"),
    (11, 5, "May"),
]
DEFAULT_EXPENSE_CATEGORIES = [
    "Transportation/Freight",
    "Gas",
    "Supplies",
    "Helpers",
    "Repair",
    "Purchase from the looses of tools",
    "Other expenses",
]
DEFAULT_ALLOCATIONS = [
    ("supplementary_feeding", "Supplementary Feeding", 35.0),
    ("school_clinic", "School Clinic", 5.0),
    ("faculty_student_development", "Faculty/Student Development", 15.0),
    ("school_operating_fund", "School Operating Fund", 25.0),
    ("he_instructional_fund", "H.E Instructional Fund", 10.0),
    ("revolving_capital_fund", "Revolving Capital Fund", 10.0),
]
EXPENSE_CELL_BY_CATEGORY = {
    "transportation/freight": "F21",
    "gas": "F22",
    "supplies": "F23",
    "helpers": "F24",
    "repair": "F25",
    "purchase from the looses of tools": "F26",
    "other expenses": "F27",
}
FUND_MONITORING_COLUMNS = ("B", "C", "D", "E", "F", "G")
EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FUTURE_FINANCIAL_REPORT_ERROR = "You cannot add a financial report for a future school year."
CURRENT_FINANCIAL_REPORT_ERROR = "Financial reports can only be saved for the current active school year."
OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OOXML_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
OOXML_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_NS = "http://www.w3.org/XML/1998/namespace"
CELL_REF_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")
ET.register_namespace("", OOXML_MAIN_NS)
ET.register_namespace("r", OOXML_REL_NS)
DEMO_BEGINNING_CASH_ON_HAND = 11834.59
DEMO_MONTHLY_REPORT_ROWS = [
    {
        "month_index": 0,
        "current_sales": 39840.00,
        "cost_of_sales": 31872.00,
        "expenses": {
            "Gas": 950.00,
            "Supplies": 1000.00,
            "Helpers": 800.00,
            "Repair": 3000.00,
        },
    },
    {
        "month_index": 1,
        "current_sales": 68483.70,
        "cost_of_sales": 54498.70,
        "expenses": {
            "Gas": 950.00,
            "Supplies": 1000.00,
            "Helpers": 1300.00,
            "Repair": 1800.00,
        },
    },
    {
        "month_index": 2,
        "current_sales": 93243.00,
        "cost_of_sales": 74595.00,
        "expenses": {
            "Gas": 1800.00,
            "Supplies": 1600.00,
            "Helpers": 2300.00,
        },
    },
    {
        "month_index": 3,
        "current_sales": 89345.00,
        "cost_of_sales": 72476.00,
        "expenses": {
            "Gas": 1900.00,
            "Supplies": 2200.00,
            "Helpers": 2100.00,
        },
    },
    {
        "month_index": 4,
        "current_sales": 70345.00,
        "cost_of_sales": 54345.00,
        "expenses": {
            "Gas": 400.00,
            "Supplies": 915.00,
            "Helpers": 2600.00,
        },
    },
    {
        "month_index": 5,
        "current_sales": 190480.00,
        "cost_of_sales": 163840.00,
        "expenses": {
            "Gas": 300.00,
            "Supplies": 1830.00,
            "Helpers": 1480.00,
            "Repair": 6500.00,
        },
    },
    {
        "month_index": 6,
        "current_sales": 85165.00,
        "cost_of_sales": 68038.00,
        "expenses": {
            "Gas": 915.00,
            "Supplies": 1800.00,
            "Helpers": 6800.00,
        },
    },
    {
        "month_index": 7,
        "current_sales": 208342.00,
        "cost_of_sales": 179482.00,
        "expenses": {
            "Transportation/Freight": 450.00,
            "Gas": 2890.00,
            "Supplies": 3650.00,
            "Helpers": 2300.00,
            "Repair": 2800.00,
        },
    },
    {
        "month_index": 8,
        "current_sales": 164280.00,
        "cost_of_sales": 132940.00,
        "expenses": {
            "Transportation/Freight": 350.00,
            "Gas": 2100.00,
            "Supplies": 2400.00,
            "Helpers": 2200.00,
            "Other expenses": 300.00,
        },
    },
    {
        "month_index": 9,
        "current_sales": 152760.00,
        "cost_of_sales": 124180.00,
        "expenses": {
            "Transportation/Freight": 300.00,
            "Gas": 1750.00,
            "Supplies": 2150.00,
            "Helpers": 2150.00,
            "Repair": 600.00,
        },
    },
    {
        "month_index": 10,
        "current_sales": 167420.00,
        "cost_of_sales": 136050.00,
        "expenses": {
            "Transportation/Freight": 320.00,
            "Gas": 1950.00,
            "Supplies": 2280.00,
            "Helpers": 2350.00,
            "Repair": 900.00,
        },
    },
    {
        "month_index": 11,
        "current_sales": 173880.00,
        "cost_of_sales": 140920.00,
        "expenses": {
            "Transportation/Freight": 400.00,
            "Gas": 2200.00,
            "Supplies": 2460.00,
            "Helpers": 2500.00,
            "Repair": 1200.00,
        },
    },
]
DEMO_MONTHLY_REPORTS_BY_INDEX = {
    int(item["month_index"]): item for item in DEMO_MONTHLY_REPORT_ROWS
}


def require_financial_report_user(
    current_user: models.User = Depends(auth.get_current_user),
) -> models.User:
    if str(current_user.role or "").strip().lower() not in FINANCIAL_REPORT_ROLES:
        raise HTTPException(status_code=403, detail="Admin or staff access required")
    return current_user


def _template_path() -> str:
    return os.path.join(os.path.abspath(os.path.dirname(__file__)), "report_templates", TEMPLATE_FILENAME)


def _format_school_year_name(start_year: int, end_year: int) -> str:
    return f"{int(start_year)}-{int(end_year)}"


def _resolve_current_active_school_year_bounds() -> tuple[int, int]:
    today = get_ph_today()
    start_year = today.year if today.month >= 6 else today.year - 1
    return start_year, start_year + 1


def _compare_school_year_to_current(start_year: int, end_year: int) -> int:
    current_start_year, current_end_year = _resolve_current_active_school_year_bounds()
    selected_bounds = (int(start_year), int(end_year))
    current_bounds = (current_start_year, current_end_year)

    if selected_bounds > current_bounds:
        return 1
    if selected_bounds < current_bounds:
        return -1
    return 0


def _validate_current_active_school_year(start_year: int, end_year: int) -> None:
    comparison = _compare_school_year_to_current(start_year, end_year)
    if comparison > 0:
        raise HTTPException(status_code=400, detail=FUTURE_FINANCIAL_REPORT_ERROR)
    if comparison < 0:
        raise HTTPException(status_code=400, detail=CURRENT_FINANCIAL_REPORT_ERROR)


def _validate_school_year_write_allowed(school_year: models.SchoolYear) -> None:
    _validate_current_active_school_year(school_year.start_year, school_year.end_year)


def _school_year_is_future(school_year: models.SchoolYear) -> bool:
    return _compare_school_year_to_current(school_year.start_year, school_year.end_year) > 0


def _round_money(value) -> float:
    return round(float(value or 0.0), 2)


def _month_calendar_year(school_year: models.SchoolYear, month_number: int) -> int:
    return school_year.start_year if month_number >= 6 else school_year.end_year


def _load_school_year(db: Session, school_year_id: int) -> Optional[models.SchoolYear]:
    return (
        db.query(models.SchoolYear)
        .options(
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.expenses),
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.SchoolYear.allocations),
        )
        .filter(models.SchoolYear.id == school_year_id)
        .first()
    )


def _create_default_expenses(db: Session, report: models.MonthlyReport) -> bool:
    existing_names = {str(expense.category or "").strip().lower() for expense in report.expenses}
    changed = False

    for sort_order, category in enumerate(DEFAULT_EXPENSE_CATEGORIES):
        if category.strip().lower() in existing_names:
            continue
        db.add(
            models.Expense(
                report_id=report.id,
                category=category,
                amount=0.0,
                sort_order=sort_order,
            )
        )
        changed = True

    return changed


def _ensure_school_year_defaults(db: Session, school_year: models.SchoolYear) -> bool:
    changed = False
    existing_allocations = {
        str(allocation.category_key or "").strip().lower(): allocation
        for allocation in school_year.allocations
    }

    for sort_order, (category_key, label, percentage) in enumerate(DEFAULT_ALLOCATIONS):
        if category_key in existing_allocations:
            continue
        db.add(
            models.Allocation(
                school_year_id=school_year.id,
                category_key=category_key,
                label=label,
                percentage=percentage,
                sort_order=sort_order,
            )
        )
        changed = True

    existing_reports = {report.month_index: report for report in school_year.monthly_reports}
    for month_index, month_number, month_name in MONTH_SEQUENCE:
        if month_index in existing_reports:
            continue

        report = models.MonthlyReport(
            school_year_id=school_year.id,
            month_index=month_index,
            month_number=month_number,
            month_name=month_name,
            calendar_year=_month_calendar_year(school_year, month_number),
        )
        db.add(report)
        db.flush()
        school_year.monthly_reports.append(report)
        changed = True

    for report in school_year.monthly_reports:
        if _create_default_expenses(db, report):
            changed = True

    if changed:
        db.flush()

    return changed


def _school_year_has_report_values(school_year: models.SchoolYear) -> bool:
    return any(
        any(
            [
                report.beginning_cash_on_hand,
                report.current_sales,
                report.other_income,
                report.purchases,
                report.inventory_used,
                report.product_cost,
                any(float(expense.amount or 0.0) for expense in report.expenses),
            ]
        )
        for report in school_year.monthly_reports
    )


def clear_financial_reporting_tables(db: Session) -> None:
    db.query(models.Expense).delete(synchronize_session=False)
    db.query(models.FundMonitoringEntry).delete(synchronize_session=False)
    db.query(models.Allocation).delete(synchronize_session=False)
    db.query(models.MonthlyReport).delete(synchronize_session=False)
    db.query(models.SchoolYear).delete(synchronize_session=False)
    db.flush()


def seed_demo_financial_reporting(db: Session, *, reset: bool = False) -> dict:
    if reset:
        clear_financial_reporting_tables(db)

    start_year, end_year = _resolve_current_active_school_year_bounds()
    school_year_name = _format_school_year_name(start_year, end_year)
    existing_school_years = int(db.query(models.SchoolYear).count())
    school_year = (
        db.query(models.SchoolYear)
        .options(
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.expenses),
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.SchoolYear.allocations),
        )
        .filter(models.SchoolYear.name == school_year_name)
        .first()
    )
    active_school_year = (
        db.query(models.SchoolYear)
        .options(
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.expenses),
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.SchoolYear.allocations),
        )
        .filter(models.SchoolYear.is_active.is_(True))
        .order_by(models.SchoolYear.updated_at.desc(), models.SchoolYear.id.desc())
        .first()
    )
    active_school_year_is_current = (
        bool(active_school_year)
        and _compare_school_year_to_current(
            active_school_year.start_year,
            active_school_year.end_year,
        )
        == 0
    )

    if not reset and active_school_year_is_current and not _school_year_has_report_values(active_school_year):
        school_year = active_school_year
        school_year_name = active_school_year.name
    elif school_year and not reset and _school_year_has_report_values(school_year):
        if (
            active_school_year_is_current
            and active_school_year.id != school_year.id
            and not _school_year_has_report_values(active_school_year)
        ):
            school_year = active_school_year
            school_year_name = active_school_year.name
        else:
            return {
                "message": f"Monthly canteen reporting already seeded for {school_year_name}.",
                "school_year": school_year_name,
                "created": False,
                "months_populated": 0,
            }

    created = False
    if not school_year:
        school_year = models.SchoolYear(
            name=school_year_name,
            start_year=start_year,
            end_year=end_year,
            is_active=existing_school_years == 0,
        )
        db.add(school_year)
        db.flush()
        created = True

    _ensure_school_year_defaults(db, school_year)
    db.flush()

    school_year = _load_school_year(db, school_year.id) or school_year
    reports_by_index = {int(report.month_index): report for report in school_year.monthly_reports}
    running_beginning_cash = _round_money(DEMO_BEGINNING_CASH_ON_HAND)

    for month_index, month_number, month_name in MONTH_SEQUENCE:
        report = reports_by_index.get(month_index)
        if not report:
            continue

        demo_row = DEMO_MONTHLY_REPORTS_BY_INDEX.get(month_index, {})
        current_sales = _round_money(demo_row.get("current_sales"))
        cost_of_sales = _round_money(demo_row.get("cost_of_sales"))
        expense_values = {
            str(category or "").strip().lower(): _round_money(amount)
            for category, amount in dict(demo_row.get("expenses") or {}).items()
        }

        report.month_number = month_number
        report.month_name = month_name
        report.calendar_year = _month_calendar_year(school_year, month_number)
        report.beginning_cash_on_hand = running_beginning_cash
        report.current_sales = current_sales
        report.other_income = 0.0
        report.purchases = 0.0
        report.inventory_used = 0.0
        report.product_cost = cost_of_sales
        report.notes = "Seeded demo monthly canteen reporting data"

        _create_default_expenses(db, report)
        expense_map = {
            str(expense.category or "").strip().lower(): expense
            for expense in report.expenses
        }

        total_operating_expenses = 0.0
        for sort_order, category in enumerate(DEFAULT_EXPENSE_CATEGORIES):
            normalized_category = category.strip().lower()
            expense = expense_map.get(normalized_category)
            if not expense:
                expense = models.Expense(
                    report_id=report.id,
                    category=category,
                    amount=0.0,
                    sort_order=sort_order,
                )
                db.add(expense)
                report.expenses.append(expense)
                expense_map[normalized_category] = expense

            amount = _round_money(expense_values.get(normalized_category))
            expense.category = category
            expense.amount = amount
            expense.sort_order = sort_order
            total_operating_expenses += amount

        running_beginning_cash = _round_money(
            running_beginning_cash + (current_sales - cost_of_sales) - total_operating_expenses
        )

    db.flush()

    return {
        "message": f"Monthly canteen reporting demo data seeded for {school_year_name}.",
        "school_year": school_year_name,
        "created": created,
        "months_populated": len(DEMO_MONTHLY_REPORT_ROWS),
    }


def _serialize_expense(expense: models.Expense) -> dict:
    return {
        "id": expense.id,
        "category": expense.category,
        "amount": _round_money(expense.amount),
        "sort_order": int(expense.sort_order or 0),
    }


def _serialize_allocation(
    allocation: models.Allocation,
    net_profit: float = 0.0,
    *,
    fund_interest: float = 0.0,
    fund_expenses: float = 0.0,
    fund_others: float = 0.0,
    fund_cash_on_bank: float = 0.0,
) -> dict:
    percentage = float(allocation.percentage or 0.0)
    return {
        "id": allocation.id,
        "category_key": allocation.category_key,
        "label": allocation.label,
        "percentage": round(percentage, 2),
        "opening_balance": _round_money(getattr(allocation, "opening_balance", 0.0)),
        "sort_order": int(allocation.sort_order or 0),
        "amount": _round_money(net_profit * percentage / 100.0),
        "fund_interest": _round_money(fund_interest),
        "fund_expenses": _round_money(fund_expenses),
        "fund_others": _round_money(fund_others),
        "fund_cash_on_bank": _round_money(fund_cash_on_bank),
    }


def _fund_entry_map(report: models.MonthlyReport) -> dict[str, models.FundMonitoringEntry]:
    return {
        str(entry.category_key or "").strip(): entry
        for entry in getattr(report, "fund_entries", []) or []
    }


def _build_report_month_bounds(report: models.MonthlyReport) -> tuple[datetime, datetime]:
    last_day = calendar.monthrange(int(report.calendar_year), int(report.month_number))[1]
    return build_ph_date_range_bounds(
        f"{int(report.calendar_year):04d}-{int(report.month_number):02d}-01",
        f"{int(report.calendar_year):04d}-{int(report.month_number):02d}-{last_day:02d}",
    )


def _get_report_transaction_sales(db: Session, report: models.MonthlyReport) -> float:
    start_utc, end_utc = _build_report_month_bounds(report)
    # A transaction row is created only after POS checkout succeeds; this system does not retain
    # pending POS transactions in the transactions table.
    total_sales = (
        db.query(func.coalesce(func.sum(models.Transaction.total), 0.0))
        .filter(models.Transaction.created_at.between(start_utc, end_utc))
        .scalar()
    )
    return _round_money(total_sales)


def _serialize_report(
    report: models.MonthlyReport,
    allocations: list[models.Allocation],
    *,
    beginning_cash_auto: Optional[float] = None,
    current_sales_auto: Optional[float] = None,
) -> dict:
    expenses = sorted(report.expenses, key=lambda item: (item.sort_order, item.id))
    serialized_expenses = [_serialize_expense(expense) for expense in expenses]
    saved_beginning_cash = _round_money(report.beginning_cash_on_hand)
    saved_current_sales = _round_money(report.current_sales)
    beginning_cash_manual_override = bool(getattr(report, "beginning_cash_manual_override", False))
    current_sales_manual_override = bool(getattr(report, "current_sales_manual_override", False))
    beginning_cash = (
        saved_beginning_cash
        if beginning_cash_manual_override or beginning_cash_auto is None
        else _round_money(beginning_cash_auto)
    )
    transaction_current_sales = (
        None if current_sales_auto is None else _round_money(current_sales_auto)
    )
    current_sales = (
        saved_current_sales
        if current_sales_manual_override or transaction_current_sales is None
        else transaction_current_sales
    )
    other_income = _round_money(report.other_income)
    purchases = _round_money(report.purchases)
    inventory_used = _round_money(report.inventory_used)
    product_cost = _round_money(report.product_cost)
    cost_of_sales = _round_money(purchases + inventory_used + product_cost)
    total_operating_expenses = _round_money(sum(expense["amount"] for expense in serialized_expenses))
    gross_income = _round_money(current_sales - cost_of_sales)
    # Other income is treated as non-sales income added after gross income.
    net_profit = _round_money(gross_income + other_income - total_operating_expenses)
    ending_cash = _round_money(beginning_cash + net_profit)
    total_expenses = _round_money(cost_of_sales + total_operating_expenses)
    fund_entries_by_key = _fund_entry_map(report)
    allocations_breakdown = []
    for allocation in sorted(allocations, key=lambda item: (item.sort_order, item.id)):
        category_key = str(allocation.category_key or "").strip()
        fund_entry = fund_entries_by_key.get(category_key)
        allocations_breakdown.append(
            _serialize_allocation(
                allocation,
                net_profit,
                fund_interest=getattr(fund_entry, "interest", 0.0) if fund_entry else 0.0,
                fund_expenses=getattr(fund_entry, "expenses", 0.0) if fund_entry else 0.0,
                fund_others=getattr(fund_entry, "others", 0.0) if fund_entry else 0.0,
                fund_cash_on_bank=getattr(fund_entry, "cash_on_bank", 0.0) if fund_entry else 0.0,
            )
        )

    return {
        "id": report.id,
        "school_year_id": report.school_year_id,
        "month_index": report.month_index,
        "month_number": report.month_number,
        "month_name": report.month_name,
        "month_short": calendar.month_abbr[int(report.month_number or 0)] or report.month_name[:3],
        "calendar_year": report.calendar_year,
        "month_label": f"{report.month_name} {report.calendar_year}",
        "beginning_cash_on_hand": beginning_cash,
        "beginning_cash_manual_override": beginning_cash_manual_override,
        "beginning_cash_auto": _round_money(beginning_cash_auto) if beginning_cash_auto is not None else None,
        "beginning_cash_source": "manual" if beginning_cash_manual_override else "automatic",
        "current_sales": current_sales,
        "current_sales_manual_override": current_sales_manual_override,
        "current_sales_auto": transaction_current_sales,
        "analytics_current_sales": transaction_current_sales,
        "transaction_current_sales": transaction_current_sales,
        "current_sales_source": "manual" if current_sales_manual_override else "automatic",
        "current_sales_locked": False,
        "other_income": other_income,
        "purchases": purchases,
        "inventory_used": inventory_used,
        "product_cost": product_cost,
        "cost_of_sales": cost_of_sales,
        "expenses": serialized_expenses,
        "total_operating_expenses": total_operating_expenses,
        "gross_income": gross_income,
        "net_profit": net_profit,
        "ending_cash": ending_cash,
        "total_expenses": total_expenses,
        "expenses_exceed_sales": total_expenses > _round_money(current_sales + other_income),
        "allocations": allocations_breakdown,
        "notes": report.notes or "",
        "updated_at": report.updated_at.isoformat() if report.updated_at else None,
        "comparison": None,
    }


def _fund_balance_total(
    allocations: list[models.Allocation],
    balances: dict[str, float],
) -> float:
    return _round_money(
        sum(
            balances.get(str(allocation.category_key or "").strip(), 0.0)
            for allocation in allocations
        )
    )


def _calculate_next_fund_balances(
    report: dict,
    allocations: list[models.Allocation],
    previous_balances: dict[str, float],
) -> dict[str, float]:
    next_balances = dict(previous_balances)
    report_allocations_by_key = {
        str(allocation.get("category_key") or "").strip(): allocation
        for allocation in report.get("allocations", [])
    }

    for allocation in allocations:
        category_key = str(allocation.category_key or "").strip()
        previous_balance = _round_money(previous_balances.get(category_key, 0.0))
        report_allocation = report_allocations_by_key.get(category_key, {})
        net_income = _round_money(report_allocation.get("amount", 0.0))
        interest = _round_money(report_allocation.get("fund_interest", 0.0))
        expenses = _round_money(report_allocation.get("fund_expenses", 0.0))
        others = _round_money(report_allocation.get("fund_others", 0.0))
        next_balances[category_key] = _round_money(
            previous_balance + interest + net_income - expenses - others
        )

    return next_balances


def _build_auto_input_payload(
    report: dict,
    *,
    previous_report: Optional[dict],
    previous_current_balance_total: Optional[float],
    transaction_sales: float,
    historical_reports: list[dict],
) -> tuple[dict, dict]:
    historical_cost_reports = [
        item for item in historical_reports if item["current_sales"] > 0 and item["cost_of_sales"] > 0
    ]
    historical_operation_reports = [
        item for item in historical_reports if item["total_operating_expenses"] > 0
    ]

    if historical_cost_reports:
        historical_cost_ratio = sum(item["cost_of_sales"] for item in historical_cost_reports) / max(
            sum(item["current_sales"] for item in historical_cost_reports),
            1,
        )
    else:
        historical_cost_ratio = None

    auto_inputs = {
        "beginning_cash_on_hand": {
            "value": _round_money(
                previous_current_balance_total
                if previous_report is not None and previous_current_balance_total is not None
                else report["beginning_cash_on_hand"]
            ),
            "source": (
                f'Auto-carried from {previous_report["month_label"]} Current Balance total'
                if previous_report is not None and previous_current_balance_total is not None
                else 'Using the saved value for this month'
            ),
        },
        "current_sales": {
            "value": transaction_sales,
            "source": (
                f'Auto-calculated from POS transactions for {report["month_label"]}'
                if transaction_sales > 0
                else 'No POS transactions found for this month yet'
            ),
        },
        "cost_of_sales": {
            "value": 0.0,
            "source": 'No historical cost-of-sales pattern found yet',
        },
        "operation_expenses": {
            "value": 0.0,
            "source": 'No historical operation-expense pattern found yet',
        },
    }

    if report["cost_of_sales"] > 0:
        auto_inputs["cost_of_sales"] = {
            "value": report["cost_of_sales"],
            "source": 'Using the saved cost of sales for this month',
        }
    elif previous_report and previous_report["cost_of_sales"] > 0:
        auto_inputs["cost_of_sales"] = {
            "value": previous_report["cost_of_sales"],
            "source": f'Copied from {previous_report["month_label"]} cost of sales',
        }
    elif historical_cost_ratio is not None and transaction_sales > 0:
        auto_inputs["cost_of_sales"] = {
            "value": _round_money(transaction_sales * historical_cost_ratio),
            "source": 'Estimated from historical cost-of-sales rate',
        }

    if report["total_operating_expenses"] > 0:
        auto_inputs["operation_expenses"] = {
            "value": report["total_operating_expenses"],
            "source": 'Using the saved operation expenses for this month',
        }
    elif previous_report and previous_report["total_operating_expenses"] > 0:
        auto_inputs["operation_expenses"] = {
            "value": previous_report["total_operating_expenses"],
            "source": f'Copied from {previous_report["month_label"]} operation expenses',
        }
    elif historical_operation_reports:
        auto_inputs["operation_expenses"] = {
            "value": _round_money(
                sum(item["total_operating_expenses"] for item in historical_operation_reports)
                / len(historical_operation_reports)
            ),
            "source": 'Estimated from average historical operation expenses',
        }

    default_inputs = {
        "beginning_cash_on_hand": auto_inputs["beginning_cash_on_hand"]["value"],
        "current_sales": report["current_sales"],
        "cost_of_sales": report["cost_of_sales"],
        "operation_expenses": report["total_operating_expenses"],
    }

    return auto_inputs, {
        key: _round_money(value) for key, value in default_inputs.items()
    }


def _build_dashboard(serialized_reports: list[dict], allocations: list[models.Allocation]) -> dict:
    total_sales = _round_money(sum(report["current_sales"] for report in serialized_reports))
    total_expenses = _round_money(sum(report["total_expenses"] for report in serialized_reports))
    total_net_profit = _round_money(sum(report["net_profit"] for report in serialized_reports))
    allocation_percent_total = round(sum(float(item.percentage or 0.0) for item in allocations), 2)

    populated_reports = [
        report
        for report in serialized_reports
        if any(
            [
                report["beginning_cash_on_hand"],
                report["current_sales"],
                report["other_income"],
                report["cost_of_sales"],
                report["total_operating_expenses"],
            ]
        )
    ]

    if populated_reports:
        best_month = max(populated_reports, key=lambda item: (item["net_profit"], item["current_sales"]))
        lowest_month = min(populated_reports, key=lambda item: (item["net_profit"], item["current_sales"]))
    else:
        best_month = None
        lowest_month = None

    return {
        "total_monthly_sales": total_sales,
        "total_expenses": total_expenses,
        "net_profit": total_net_profit,
        "best_month": (
            {
                "label": best_month["month_label"],
                "net_profit": best_month["net_profit"],
                "sales": best_month["current_sales"],
            }
            if best_month
            else None
        ),
        "lowest_month": (
            {
                "label": lowest_month["month_label"],
                "net_profit": lowest_month["net_profit"],
                "sales": lowest_month["current_sales"],
            }
            if lowest_month
            else None
        ),
        "warning_count": sum(1 for report in serialized_reports if report["expenses_exceed_sales"]),
        "allocation_percent_total": allocation_percent_total,
        "monthly_sales_chart": [
            {
                "label": report["month_short"],
                "sales": report["current_sales"],
                "expenses": report["total_expenses"],
            }
            for report in serialized_reports
        ],
        "monthly_profit_chart": [
            {
                "label": report["month_short"],
                "net_profit": report["net_profit"],
                "ending_cash": report["ending_cash"],
            }
            for report in serialized_reports
        ],
    }


def _serialize_school_year_detail(db: Session, school_year: models.SchoolYear) -> dict:
    allocations = sorted(school_year.allocations, key=lambda item: (item.sort_order, item.id))
    reports = sorted(school_year.monthly_reports, key=lambda item: item.month_index)
    transaction_sales_by_report_id = {
        report.id: _get_report_transaction_sales(db, report)
        for report in reports
    }
    beginning_cash_carry_forward = _get_beginning_cash_carry_forward(db, school_year)

    previous_report = None
    previous_fund_balances: dict[str, float] = {
        str(allocation.category_key or "").strip(): _round_money(
            getattr(allocation, "opening_balance", 0.0)
        )
        for allocation in allocations
    }
    historical_reports = []
    serialized_reports = []
    for monthly_report in reports:
        transaction_sales = transaction_sales_by_report_id[monthly_report.id]
        previous_balance_total = _fund_balance_total(allocations, previous_fund_balances)
        if previous_report is not None:
            beginning_cash_auto = previous_balance_total
            beginning_cash_auto_source = (
                f'Auto-carried from {previous_report["month_label"]} Current Balance total'
            )
        elif beginning_cash_carry_forward:
            beginning_cash_auto = _round_money(beginning_cash_carry_forward["amount"])
            beginning_cash_auto_source = (
                f'Auto-carried from {beginning_cash_carry_forward["source_month_label"]} '
                f'Current Balance total'
            )
        else:
            beginning_cash_auto = 0.0
            beginning_cash_auto_source = 'No previous report found; defaulted to PHP 0.00'

        report = _serialize_report(
            monthly_report,
            allocations,
            beginning_cash_auto=beginning_cash_auto,
            current_sales_auto=transaction_sales,
        )
        auto_inputs, default_inputs = _build_auto_input_payload(
            report,
            previous_report=previous_report,
            previous_current_balance_total=previous_balance_total if previous_report else None,
            transaction_sales=transaction_sales,
            historical_reports=historical_reports,
        )
        auto_inputs["beginning_cash_on_hand"] = {
            "value": beginning_cash_auto,
            "source": beginning_cash_auto_source,
        }
        default_inputs["beginning_cash_on_hand"] = report["beginning_cash_on_hand"]
        default_inputs["current_sales"] = report["current_sales"]
        report["auto_inputs"] = auto_inputs
        report["default_inputs"] = default_inputs
        report["auto_fill_applied_by_default"] = not (
            report["beginning_cash_manual_override"] and report["current_sales_manual_override"]
        )
        report["fund_previous_balance_total"] = previous_balance_total
        report["beginning_cash_locked"] = False
        report["beginning_cash_source"] = (
            "manual" if report["beginning_cash_manual_override"] else "automatic"
        )
        report["beginning_cash_carry_forward"] = beginning_cash_carry_forward

        if previous_report:
            report["comparison"] = {
                "previous_month_label": previous_report["month_label"],
                "sales_delta": _round_money(report["current_sales"] - previous_report["current_sales"]),
                "net_profit_delta": _round_money(report["net_profit"] - previous_report["net_profit"]),
            }

        next_fund_balances = _calculate_next_fund_balances(
            report,
            allocations,
            previous_fund_balances,
        )
        report["fund_current_balance_total"] = _fund_balance_total(allocations, next_fund_balances)
        previous_fund_balances = next_fund_balances
        previous_report = report
        historical_reports.append(report)
        serialized_reports.append(report)

    return {
        "school_year": {
            "id": school_year.id,
            "name": school_year.name,
            "start_year": school_year.start_year,
            "end_year": school_year.end_year,
            "is_active": bool(school_year.is_active),
            "created_at": school_year.created_at.isoformat() if school_year.created_at else None,
            "updated_at": school_year.updated_at.isoformat() if school_year.updated_at else None,
        },
        "allocations": [_serialize_allocation(allocation) for allocation in allocations],
        "reports": serialized_reports,
        "dashboard": _build_dashboard(serialized_reports, allocations),
    }


def _load_previous_school_year(
    db: Session,
    school_year: models.SchoolYear,
) -> Optional[models.SchoolYear]:
    previous_school_year = (
        db.query(models.SchoolYear)
        .options(
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.expenses),
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.SchoolYear.allocations),
        )
        .filter(
            models.SchoolYear.start_year == int(school_year.start_year) - 1,
            models.SchoolYear.end_year == int(school_year.start_year),
        )
        .first()
    )
    if previous_school_year:
        return previous_school_year

    return (
        db.query(models.SchoolYear)
        .options(
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.expenses),
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.SchoolYear.allocations),
        )
        .filter(
            models.SchoolYear.id != school_year.id,
            models.SchoolYear.end_year <= int(school_year.start_year),
        )
        .order_by(
            models.SchoolYear.end_year.desc(),
            models.SchoolYear.start_year.desc(),
            models.SchoolYear.id.desc(),
        )
        .first()
    )


def _calculate_school_year_final_current_balance(
    db: Session,
    school_year: models.SchoolYear,
) -> Optional[dict]:
    reports = sorted(school_year.monthly_reports, key=lambda item: item.month_index)
    if not reports:
        return None

    allocations = sorted(school_year.allocations, key=lambda item: (item.sort_order, item.id))
    last_report_payload = None

    if not allocations:
        for report in reports:
            last_report_payload = _serialize_report(
                report,
                allocations,
                current_sales_auto=_get_report_transaction_sales(db, report),
            )
        if not last_report_payload:
            return None
        return {
            "amount": _round_money(last_report_payload["ending_cash"]),
            "month_label": last_report_payload["month_label"],
            "school_year_name": school_year.name,
        }

    previous_fund_balances: dict[str, float] = {
        str(allocation.category_key or "").strip(): _round_money(
            getattr(allocation, "opening_balance", 0.0)
        )
        for allocation in allocations
    }
    final_balance = 0.0

    for report in reports:
        report_payload = _serialize_report(
            report,
            allocations,
            current_sales_auto=_get_report_transaction_sales(db, report),
        )
        previous_fund_balances = _calculate_next_fund_balances(
            report_payload,
            allocations,
            previous_fund_balances,
        )
        final_balance = _fund_balance_total(allocations, previous_fund_balances)
        last_report_payload = report_payload

    if not last_report_payload:
        return None

    return {
        "amount": _round_money(final_balance),
        "month_label": last_report_payload["month_label"],
        "school_year_name": school_year.name,
    }


def _get_beginning_cash_carry_forward(
    db: Session,
    school_year: models.SchoolYear,
) -> Optional[dict]:
    previous_school_year = _load_previous_school_year(db, school_year)
    if not previous_school_year:
        return None

    carry_forward = _calculate_school_year_final_current_balance(db, previous_school_year)
    if not carry_forward:
        return None

    return {
        "amount": _round_money(carry_forward["amount"]),
        "source_school_year_name": carry_forward["school_year_name"],
        "source_month_label": carry_forward["month_label"],
    }


def _apply_beginning_cash_carry_forward(
    db: Session,
    school_year: models.SchoolYear,
) -> bool:
    carry_forward = _get_beginning_cash_carry_forward(db, school_year)
    if not carry_forward:
        return False

    first_report = next(
        (report for report in school_year.monthly_reports if int(report.month_index or 0) == 0),
        None,
    )
    if not first_report:
        return False

    carry_forward_amount = _round_money(carry_forward["amount"])
    if _round_money(first_report.beginning_cash_on_hand) == carry_forward_amount:
        return False

    first_report.beginning_cash_on_hand = carry_forward_amount
    return True


def _get_report_beginning_cash_carry_forward(
    db: Session,
    report: models.MonthlyReport,
) -> Optional[dict]:
    if int(report.month_index or 0) != 0 or not report.school_year:
        return None

    return _get_beginning_cash_carry_forward(db, report.school_year)


def _build_school_year_summary(db: Session, school_year: models.SchoolYear) -> dict:
    allocations = sorted(school_year.allocations, key=lambda item: (item.sort_order, item.id))
    reports = sorted(school_year.monthly_reports, key=lambda item: item.month_index)
    serialized_reports = [
        _serialize_report(
            report,
            allocations,
            current_sales_auto=_get_report_transaction_sales(db, report),
        )
        for report in reports
    ]
    dashboard = _build_dashboard(serialized_reports, allocations)
    months_with_entries = sum(
        1
        for report in serialized_reports
        if any(
            [
                report["beginning_cash_on_hand"],
                report["current_sales"],
                report["other_income"],
                report["cost_of_sales"],
                report["total_operating_expenses"],
            ]
        )
    )
    final_balance = _calculate_school_year_final_current_balance(db, school_year)
    opening_beginning_cash = (
        _round_money(serialized_reports[0]["beginning_cash_on_hand"])
        if serialized_reports
        else 0.0
    )
    ending_balance = (
        _round_money(final_balance["amount"])
        if final_balance
        else _round_money(serialized_reports[-1]["ending_cash"] if serialized_reports else 0.0)
    )

    return {
        "id": school_year.id,
        "name": school_year.name,
        "start_year": school_year.start_year,
        "end_year": school_year.end_year,
        "is_active": bool(school_year.is_active),
        "status": "Active" if school_year.is_active else "Closed",
        "opening_beginning_cash": opening_beginning_cash,
        "ending_balance": ending_balance,
        "months_with_entries": months_with_entries,
        "report_count": len(serialized_reports),
        "total_sales": dashboard["total_monthly_sales"],
        "net_profit": dashboard["net_profit"],
        "updated_at": school_year.updated_at.isoformat() if school_year.updated_at else None,
    }


def _ensure_and_reload_school_year(db: Session, school_year_id: int) -> models.SchoolYear:
    school_year = _load_school_year(db, school_year_id)
    if not school_year:
        raise HTTPException(status_code=404, detail="School year not found")

    if not _school_year_is_future(school_year) and _ensure_school_year_defaults(db, school_year):
        db.commit()
        school_year = _load_school_year(db, school_year_id)

    return school_year


def _audit_log(
    db: Session,
    *,
    user_id: int,
    action: str,
    details: str,
    user_type: Optional[str] = None,
    request: Optional[Request] = None,
) -> None:
    resolved_user_type = user_type
    if not resolved_user_type:
        if user_id:
            user = db.query(models.User).filter(models.User.id == user_id).first()
            if user and user.role:
                resolved_user_type = user.role
            else:
                resolved_user_type = "user"
        else:
            resolved_user_type = "system"

    db.add(
        models.AuditLog(
            user_id=user_id,
            user_type=resolved_user_type,
            action=action,
            details=details,
            ip_address=request.client.host if request and request.client else None,
        )
    )


def _sqlite_database_path() -> Optional[str]:
    if not SQLALCHEMY_DATABASE_URL.startswith("sqlite:///"):
        return None
    return os.path.abspath(SQLALCHEMY_DATABASE_URL.replace("sqlite:///", "", 1))


def _remove_file_if_exists(path: str) -> None:
    try:
        if path and os.path.isfile(path):
            os.remove(path)
    except OSError:
        pass


def _ooxml_name(local_name: str) -> str:
    return f"{{{OOXML_MAIN_NS}}}{local_name}"


def _ooxml_rel_name(local_name: str) -> str:
    return f"{{{OOXML_REL_NS}}}{local_name}"


def _package_rel_name(local_name: str) -> str:
    return f"{{{OOXML_PACKAGE_REL_NS}}}{local_name}"


def _serialize_ooxml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _column_letters_to_index(letters: str) -> int:
    column_index = 0
    for letter in letters:
        column_index = column_index * 26 + (ord(letter) - ord("A") + 1)
    return column_index


def _split_cell_ref(cell_ref: str) -> tuple[str, int]:
    match = CELL_REF_RE.match(str(cell_ref or "").upper())
    if not match:
        raise ValueError(f"Invalid Excel cell reference: {cell_ref}")
    return match.group(1), int(match.group(2))


def _cell_sort_key(cell_ref: str) -> tuple[int, int]:
    column_letters, row_number = _split_cell_ref(cell_ref)
    return row_number, _column_letters_to_index(column_letters)


def _find_or_create_row(sheet_data: ET.Element, row_number: int) -> ET.Element:
    for row in sheet_data.findall(_ooxml_name("row")):
        if int(row.attrib.get("r", "0") or 0) == row_number:
            return row

    row_element = ET.Element(_ooxml_name("row"), {"r": str(row_number)})
    rows = list(sheet_data.findall(_ooxml_name("row")))
    insert_at = len(rows)
    for index, row in enumerate(rows):
        if int(row.attrib.get("r", "0") or 0) > row_number:
            insert_at = index
            break
    sheet_data.insert(insert_at, row_element)
    return row_element


def _find_or_create_cell(root: ET.Element, cell_ref: str) -> ET.Element:
    normalized_ref = str(cell_ref or "").upper()
    column_letters, row_number = _split_cell_ref(normalized_ref)
    column_index = _column_letters_to_index(column_letters)
    sheet_data = root.find(_ooxml_name("sheetData"))
    if sheet_data is None:
        sheet_data = ET.SubElement(root, _ooxml_name("sheetData"))

    row = _find_or_create_row(sheet_data, row_number)
    for cell in row.findall(_ooxml_name("c")):
        if cell.attrib.get("r") == normalized_ref:
            return cell

    cell = ET.Element(_ooxml_name("c"), {"r": normalized_ref})
    cells = list(row.findall(_ooxml_name("c")))
    insert_at = len(cells)
    for index, existing_cell in enumerate(cells):
        existing_ref = existing_cell.attrib.get("r", "")
        if existing_ref and _cell_sort_key(existing_ref)[1] > column_index:
            insert_at = index
            break
    row.insert(insert_at, cell)
    return cell


def _remove_cell_children(cell: ET.Element, names: set[str]) -> None:
    for child in list(cell):
        if child.tag in names:
            cell.remove(child)


def _set_ooxml_number_cell(
    root: ET.Element,
    cell_ref: str,
    value,
    *,
    overwrite_formula: bool = False,
) -> bool:
    cell = _find_or_create_cell(root, cell_ref)
    if cell.find(_ooxml_name("f")) is not None and not overwrite_formula:
        return False

    _remove_cell_children(
        cell,
        {_ooxml_name("f"), _ooxml_name("v"), _ooxml_name("is")},
    )
    cell.attrib.pop("t", None)
    value_element = ET.SubElement(cell, _ooxml_name("v"))
    value_element.text = str(_round_money(value))
    return True


def _set_ooxml_text_cell(
    root: ET.Element,
    cell_ref: str,
    value: str,
    *,
    overwrite_formula: bool = False,
) -> bool:
    cell = _find_or_create_cell(root, cell_ref)
    if cell.find(_ooxml_name("f")) is not None and not overwrite_formula:
        return False

    _remove_cell_children(
        cell,
        {_ooxml_name("f"), _ooxml_name("v"), _ooxml_name("is")},
    )
    cell.set("t", "inlineStr")
    inline_string = ET.SubElement(cell, _ooxml_name("is"))
    text = ET.SubElement(inline_string, _ooxml_name("t"))
    text_value = str(value or "")
    if text_value != text_value.strip():
        text.set(f"{{{XML_NS}}}space", "preserve")
    text.text = text_value
    return True


def _set_ooxml_cell(
    root: ET.Element,
    cell_ref: str,
    value,
    *,
    value_type: str = "number",
    overwrite_formula: bool = False,
) -> bool:
    if value_type == "text":
        return _set_ooxml_text_cell(
            root,
            cell_ref,
            str(value or ""),
            overwrite_formula=overwrite_formula,
        )
    return _set_ooxml_number_cell(
        root,
        cell_ref,
        value,
        overwrite_formula=overwrite_formula,
    )


def _normalize_workbook_target(target: str) -> str:
    target = str(target or "").replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _sheet_paths_by_name(workbook_archive: ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(workbook_archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(workbook_archive.read("xl/_rels/workbook.xml.rels"))
    relationship_targets = {
        relationship.attrib.get("Id"): _normalize_workbook_target(
            relationship.attrib.get("Target", "")
        )
        for relationship in rels_root.findall(_package_rel_name("Relationship"))
    }

    sheets = workbook_root.find(_ooxml_name("sheets"))
    if sheets is None:
        return {}

    return {
        sheet.attrib.get("name", ""): relationship_targets.get(
            sheet.attrib.get(_ooxml_rel_name("id")),
            "",
        )
        for sheet in sheets.findall(_ooxml_name("sheet"))
        if sheet.attrib.get("name")
    }


def _prepare_workbook_xml_for_recalculation(
    workbook_xml: bytes,
    *,
    active_sheet_name: Optional[str] = None,
) -> bytes:
    root = ET.fromstring(workbook_xml)

    if active_sheet_name:
        sheets = root.find(_ooxml_name("sheets"))
        sheet_names = [
            sheet.attrib.get("name", "")
            for sheet in sheets.findall(_ooxml_name("sheet"))
        ] if sheets is not None else []
        if active_sheet_name in sheet_names:
            active_index = sheet_names.index(active_sheet_name)
            book_views = root.find(_ooxml_name("bookViews"))
            workbook_view = (
                book_views.find(_ooxml_name("workbookView"))
                if book_views is not None
                else None
            )
            if workbook_view is not None:
                workbook_view.set("activeTab", str(active_index))

    calc_pr = root.find(_ooxml_name("calcPr"))
    if calc_pr is None:
        calc_pr = ET.SubElement(root, _ooxml_name("calcPr"))

    calc_pr.set("calcMode", "auto")
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")
    return _serialize_ooxml(root)


def _set_sheet_tab_selected(root: ET.Element, selected: bool) -> None:
    sheet_views = root.find(_ooxml_name("sheetViews"))
    if sheet_views is None:
        return

    for sheet_view in sheet_views.findall(_ooxml_name("sheetView")):
        if selected:
            sheet_view.set("tabSelected", "1")
        else:
            sheet_view.attrib.pop("tabSelected", None)


def _build_report_cell_updates(
    report: models.MonthlyReport,
    *,
    financial_values: Optional[dict] = None,
) -> tuple[list[tuple[str, object, str]], dict]:
    financial_values = financial_values or {}
    current_sales = _round_money(financial_values.get("current_sales", report.current_sales))
    beginning_cash = _round_money(
        financial_values.get("beginning_cash_on_hand", report.beginning_cash_on_hand)
    )
    cost_of_sales = _round_money(
        financial_values.get(
            "cost_of_sales",
            _round_money(report.purchases)
            + _round_money(report.inventory_used)
            + _round_money(report.product_cost),
        )
    )

    updates: list[tuple[str, object, str]] = [
        ("A13", f"For the Month of {report.month_name} {report.calendar_year}", "text"),
        ("C15", beginning_cash, "number"),
        ("F16", current_sales, "number"),
        ("F17", cost_of_sales, "number"),
        ("G32", 0.0, "number"),
        ("G33", 0.0, "number"),
        ("G34", _round_money(report.other_income), "number"),
    ]

    for cell_address in EXPENSE_CELL_BY_CATEGORY.values():
        updates.append((cell_address, 0.0, "number"))

    total_operating_expenses = 0.0
    for expense in sorted(report.expenses, key=lambda item: (item.sort_order, item.id)):
        category = str(expense.category or "").strip().lower()
        cell_address = EXPENSE_CELL_BY_CATEGORY.get(category)
        if cell_address:
            amount = _round_money(expense.amount)
            updates.append((cell_address, amount, "number"))
            total_operating_expenses += amount

    total_operating_expenses = _round_money(total_operating_expenses)
    additional_income = _round_money(report.other_income)
    net_profit = _round_money(current_sales - cost_of_sales - total_operating_expenses + additional_income)
    ending_cash = _round_money(beginning_cash + net_profit)

    return updates, {
        "beginning_cash": beginning_cash,
        "current_sales": current_sales,
        "cost_of_sales": cost_of_sales,
        "total_operating_expenses": total_operating_expenses,
        "additional_income": additional_income,
        "net_profit": net_profit,
        "ending_cash": ending_cash,
    }


def _format_allocation_header(allocation: models.Allocation) -> str:
    label = str(allocation.label or "Fund").strip().upper()
    percentage = round(float(allocation.percentage or 0.0), 2)
    return f"{label} {percentage:.2f}%"


def _build_fund_monitoring_cell_updates(
    allocations: list[models.Allocation],
    *,
    net_profit: float,
    previous_balances: dict[str, float],
    fund_entries_by_key: dict[str, models.FundMonitoringEntry],
) -> tuple[list[tuple[str, object, str]], dict[str, float]]:
    next_balances = dict(previous_balances)
    updates: list[tuple[str, object, str]] = []

    for column_letter, allocation in zip(FUND_MONITORING_COLUMNS, allocations):
        category_key = str(allocation.category_key or "").strip()
        percentage = round(float(allocation.percentage or 0.0), 2)
        previous_balance = _round_money(previous_balances.get(category_key, 0.0))
        fund_entry = fund_entries_by_key.get(category_key)
        interest = _round_money(getattr(fund_entry, "interest", 0.0) if fund_entry else 0.0)
        net_income = _round_money(_round_money(net_profit) * percentage / 100.0)
        expenses = _round_money(getattr(fund_entry, "expenses", 0.0) if fund_entry else 0.0)
        others = _round_money(getattr(fund_entry, "others", 0.0) if fund_entry else 0.0)
        cash_on_bank = _round_money(getattr(fund_entry, "cash_on_bank", 0.0) if fund_entry else 0.0)
        total_current_expenses = _round_money(expenses + others)
        current_balance = _round_money(
            previous_balance + interest + net_income - total_current_expenses
        )

        updates.extend(
            [
                (f"{column_letter}38", _format_allocation_header(allocation), "text"),
                (f"{column_letter}39", previous_balance, "number"),
                (f"{column_letter}40", interest, "number"),
                (f"{column_letter}41", net_income, "number"),
                (f"{column_letter}42", expenses, "number"),
                (f"{column_letter}43", others, "number"),
                (f"{column_letter}44", total_current_expenses, "number"),
                (f"{column_letter}45", current_balance, "number"),
                (f"{column_letter}46", cash_on_bank, "number"),
            ]
        )

        next_balances[category_key] = current_balance

    for column_letter in FUND_MONITORING_COLUMNS[len(allocations):]:
        for row_number in range(39, 47):
            updates.append((f"{column_letter}{row_number}", 0.0, "number"))

    updates.append(
        (
            "H45",
            _round_money(
                sum(
                    next_balances.get(str(allocation.category_key or "").strip(), 0.0)
                    for allocation in allocations
                )
            ),
            "number",
        )
    )
    return updates, next_balances


def _preserve_template_drawings_and_media(export_path: str, template_path: str) -> None:
    if not os.path.isfile(template_path) or not os.path.isfile(export_path):
        return

    temp_zip_path = export_path + ".tmp"
    try:
        with ZipFile(template_path, "r") as tmpl_zip:
            tmpl_names = tmpl_zip.namelist()
            media_files = {f: tmpl_zip.read(f) for f in tmpl_names if f.startswith("xl/media/")}
            drawing_files = {f: tmpl_zip.read(f) for f in tmpl_names if f.startswith("xl/drawings/")}

            template_drawing_xml = None
            for name in tmpl_names:
                if name.startswith("xl/drawings/drawing") and name.endswith(".xml") and not name.startswith("xl/drawings/_rels/"):
                    template_drawing_xml = tmpl_zip.read(name)
                    break

            template_drawing_rel = None
            for name in tmpl_names:
                if name.startswith("xl/drawings/_rels/drawing") and name.endswith(".xml.rels"):
                    template_drawing_rel = tmpl_zip.read(name)
                    break

        with ZipFile(export_path, "r") as exp_zip:
            exp_names = exp_zip.namelist()
            sheet_files = [f for f in exp_names if f.startswith("xl/worksheets/sheet") and f.endswith(".xml")]

        has_drawings = bool(media_files or drawing_files or template_drawing_xml)

        with ZipFile(export_path, "r") as exp_zip, ZipFile(temp_zip_path, "w", compression=ZIP_DEFLATED) as new_zip:
            for item in exp_zip.infolist():
                if (
                    item.filename.startswith("xl/media/")
                    or item.filename.startswith("xl/drawings/")
                    or item.filename.startswith("xl/worksheets/_rels/")
                ):
                    continue

                content = exp_zip.read(item.filename)

                if item.filename in sheet_files:
                    if has_drawings:
                        content_str = content.decode("utf-8")
                        if "<drawing" not in content_str:
                            if "</worksheet>" in content_str:
                                content_str = content_str.replace("</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                                content = content_str.encode("utf-8")
                    else:
                        content_str = content.decode("utf-8")
                        if "<drawing" in content_str:
                            content_str = re.sub(r'<drawing\s+r:id="[^"]+"\s*/>', "", content_str)
                            content = content_str.encode("utf-8")

                if item.filename == "[Content_Types].xml" and has_drawings:
                    content_str = content.decode("utf-8")
                    if 'Extension="jpg"' not in content_str:
                        if '<Default Extension="jpeg"' in content_str:
                            content_str = content_str.replace(
                                '<Default Extension="jpeg"',
                                '<Default Extension="jpg" ContentType="image/jpeg"/><Default Extension="jpeg"'
                            )
                        else:
                            content_str = content_str.replace(
                                "</Types>",
                                '<Default Extension="jpg" ContentType="image/jpeg"/></Types>'
                            )

                    drawing_overrides = ""
                    for s_file in sheet_files:
                        idx = s_file.replace("xl/worksheets/sheet", "").replace(".xml", "")
                        drawing_path = f"xl/drawings/drawing{idx}.xml"
                        if f'PartName="/{drawing_path}"' not in content_str:
                            drawing_overrides += f'<Override PartName="/{drawing_path}" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'

                    if drawing_overrides:
                        content_str = content_str.replace("</Types>", f"{drawing_overrides}</Types>")
                    content = content_str.encode("utf-8")

                new_zip.writestr(item, content)

            for filename, data in media_files.items():
                new_zip.writestr(filename, data)

            if has_drawings and template_drawing_xml:
                default_sheet_rel = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing{idx}.xml"/>'
                    '</Relationships>'
                )

                for s_file in sheet_files:
                    idx = s_file.replace("xl/worksheets/sheet", "").replace(".xml", "")

                    sheet_rel_path = f"xl/worksheets/_rels/sheet{idx}.xml.rels"
                    sheet_rel_content = default_sheet_rel.format(idx=idx).encode("utf-8")
                    new_zip.writestr(sheet_rel_path, sheet_rel_content)

                    drawing_path = f"xl/drawings/drawing{idx}.xml"
                    new_zip.writestr(drawing_path, template_drawing_xml)

                    if template_drawing_rel:
                        drawing_rel_path = f"xl/drawings/_rels/drawing{idx}.xml.rels"
                        new_zip.writestr(drawing_rel_path, template_drawing_rel)

        os.replace(temp_zip_path, export_path)
    except Exception as exc:
        print(f"Template media preservation skipped: {exc}")
        if os.path.exists(temp_zip_path):
            try:
                os.remove(temp_zip_path)
            except Exception:
                pass


def _build_school_year_workbook_export(
    db: Session,
    school_year: models.SchoolYear,
    *,
    selected_report_id: Optional[int] = None,
) -> str:
    template_path = _template_path()
    if not os.path.isfile(template_path):
        raise HTTPException(status_code=404, detail="Report template file not found")

    active_sheet_name = None
    if selected_report_id is not None:
        selected_report = next(
            (report for report in school_year.monthly_reports if report.id == selected_report_id),
            None,
        )
        if not selected_report:
            raise HTTPException(status_code=404, detail="Selected monthly report not found")
        active_sheet_name = selected_report.month_name

    allocations = sorted(school_year.allocations, key=lambda item: (item.sort_order, item.id))[
        :len(FUND_MONITORING_COLUMNS)
    ]
    effective_reports_by_id = {
        report["id"]: report
        for report in _serialize_school_year_detail(db, school_year).get("reports", [])
    }

    export_file = tempfile.NamedTemporaryFile(
        delete=False,
        prefix=f"canteen-report-{school_year.name}-",
        suffix=".xlsx",
    )
    export_path = export_file.name
    export_file.close()

    try:
        workbook = load_workbook(template_path, data_only=False)
        worksheet_names = workbook.sheetnames
        if active_sheet_name and active_sheet_name in worksheet_names:
            workbook.active = worksheet_names.index(active_sheet_name)

        previous_fund_balances: dict[str, float] = {
            str(allocation.category_key or "").strip(): _round_money(
                getattr(allocation, "opening_balance", 0.0)
            )
            for allocation in allocations
        }

        for report in sorted(school_year.monthly_reports, key=lambda item: item.month_index):
            if report.month_name not in workbook.sheetnames:
                continue

            worksheet = workbook[report.month_name]
            report_updates, result = _build_report_cell_updates(
                report,
                financial_values=effective_reports_by_id.get(report.id),
            )
            fund_updates, previous_fund_balances = _build_fund_monitoring_cell_updates(
                allocations,
                net_profit=result["net_profit"],
                previous_balances=previous_fund_balances,
                fund_entries_by_key=_fund_entry_map(report),
            )

            for cell_ref, value, value_type in [*report_updates, *fund_updates]:
                cell = worksheet[cell_ref]
                if cell.data_type == "f":
                    continue
                cell.value = value

        for worksheet in workbook.worksheets:
            worksheet.sheet_view.tabSelected = worksheet.title == active_sheet_name

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(export_path)
        _preserve_template_drawings_and_media(export_path, template_path)
    except Exception:
        _remove_file_if_exists(export_path)
        raise

    return export_path


@router.get("/api/financial-reports/school-years")
def list_school_years(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_financial_report_user),
):
    school_years = (
        db.query(models.SchoolYear)
        .options(
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.expenses),
            joinedload(models.SchoolYear.monthly_reports).joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.SchoolYear.allocations),
        )
        .order_by(models.SchoolYear.start_year.desc(), models.SchoolYear.id.desc())
        .all()
    )

    summaries = []
    for school_year in school_years:
        if not _school_year_is_future(school_year) and _ensure_school_year_defaults(db, school_year):
            db.commit()
            school_year = _load_school_year(db, school_year.id)
        summaries.append(_build_school_year_summary(db, school_year))

    return summaries


@router.post("/api/financial-reports/school-years")
def create_school_year(
    payload: schemas.FinancialSchoolYearCreate,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    start_year = int(payload.start_year)
    end_year = int(payload.end_year or (start_year + 1))

    if end_year <= start_year:
        raise HTTPException(status_code=400, detail="End year must be after the start year")

    _validate_current_active_school_year(start_year, end_year)

    school_year_name = _format_school_year_name(start_year, end_year)
    existing = db.query(models.SchoolYear).filter(models.SchoolYear.name == school_year_name).first()
    if existing:
        raise HTTPException(status_code=409, detail="School year already exists")

    if payload.set_active:
        db.query(models.SchoolYear).update({models.SchoolYear.is_active: False})

    school_year = models.SchoolYear(
        name=school_year_name,
        start_year=start_year,
        end_year=end_year,
        is_active=bool(payload.set_active),
    )
    db.add(school_year)
    db.flush()
    _ensure_school_year_defaults(db, school_year)
    _apply_beginning_cash_carry_forward(db, school_year)
    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_SCHOOL_YEAR_CREATED",
        details=f"Created school year {school_year_name}",
        request=request,
    )
    db.commit()

    return _serialize_school_year_detail(db, _ensure_and_reload_school_year(db, school_year.id))


@router.put("/api/financial-reports/school-years/{school_year_id}")
def update_school_year(
    school_year_id: int,
    payload: schemas.FinancialSchoolYearUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    school_year = _load_school_year(db, school_year_id)
    if not school_year:
        raise HTTPException(status_code=404, detail="School year not found")

    updates = payload.model_dump(exclude_unset=True)
    next_start_year = int(updates.get("start_year", school_year.start_year))
    next_end_year = int(updates.get("end_year", school_year.end_year))
    if next_end_year <= next_start_year:
        raise HTTPException(status_code=400, detail="End year must be after the start year")

    next_name = _format_school_year_name(next_start_year, next_end_year)
    duplicate = (
        db.query(models.SchoolYear)
        .filter(
            models.SchoolYear.id != school_year.id,
            models.SchoolYear.name == next_name,
        )
        .first()
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="School year already exists")

    if updates.get("is_active") is False and school_year.is_active:
        raise HTTPException(
            status_code=400,
            detail="Activate another school year before archiving the active school year.",
        )

    if updates.get("is_active") is True:
        db.query(models.SchoolYear).update({models.SchoolYear.is_active: False})

    school_year.start_year = next_start_year
    school_year.end_year = next_end_year
    school_year.name = next_name
    if "is_active" in updates:
        school_year.is_active = bool(updates["is_active"])

    for report in school_year.monthly_reports:
        report.calendar_year = _month_calendar_year(school_year, int(report.month_number or 1))

    _ensure_school_year_defaults(db, school_year)
    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_SCHOOL_YEAR_UPDATED",
        details=f"Updated school year {next_name}",
        request=request,
    )
    db.commit()

    return _serialize_school_year_detail(db, _ensure_and_reload_school_year(db, school_year.id))


def _set_school_year_active_state(
    school_year_id: int,
    *,
    is_active: bool,
    request: Request,
    db: Session,
    current: models.User,
) -> dict:
    school_year = _load_school_year(db, school_year_id)
    if not school_year:
        raise HTTPException(status_code=404, detail="School year not found")

    if is_active:
        db.query(models.SchoolYear).update({models.SchoolYear.is_active: False})
        school_year.is_active = True
        action = "FINANCIAL_REPORT_SCHOOL_YEAR_ACTIVATED"
        verb = "activated"
    else:
        if school_year.is_active:
            raise HTTPException(
                status_code=400,
                detail="Activate another school year before archiving the active school year.",
            )
        school_year.is_active = False
        action = "FINANCIAL_REPORT_SCHOOL_YEAR_ARCHIVED"
        verb = "archived"

    _audit_log(
        db,
        user_id=current.id,
        action=action,
        details=f"School year {school_year.name} {verb}",
        request=request,
    )
    db.commit()

    detail = _serialize_school_year_detail(db, _ensure_and_reload_school_year(db, school_year.id))
    detail["message"] = f"School year {detail['school_year']['name']} {verb}."
    return detail


@router.post("/api/financial-reports/school-years/{school_year_id}/activate")
def activate_school_year(
    school_year_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    return _set_school_year_active_state(
        school_year_id,
        is_active=True,
        request=request,
        db=db,
        current=current,
    )


@router.post("/api/financial-reports/school-years/{school_year_id}/archive")
def archive_school_year(
    school_year_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    return _set_school_year_active_state(
        school_year_id,
        is_active=False,
        request=request,
        db=db,
        current=current,
    )


@router.delete("/api/financial-reports/school-years/{school_year_id}")
def delete_school_year(
    school_year_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    school_year = _load_school_year(db, school_year_id)
    if not school_year:
        raise HTTPException(status_code=404, detail="School year not found")

    deleted_name = school_year.name
    db.delete(school_year)
    db.flush()

    active_school_year = (
        db.query(models.SchoolYear)
        .filter(models.SchoolYear.is_active.is_(True))
        .order_by(models.SchoolYear.start_year.desc(), models.SchoolYear.id.desc())
        .first()
    )
    if not active_school_year:
        active_school_year = (
            db.query(models.SchoolYear)
            .order_by(models.SchoolYear.start_year.desc(), models.SchoolYear.id.desc())
            .first()
        )
        if active_school_year:
            active_school_year.is_active = True

    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_SCHOOL_YEAR_DELETED",
        details=f"Deleted school year {deleted_name}",
        request=request,
    )
    db.commit()

    return {
        "message": f"School year {deleted_name} removed.",
        "deleted_id": school_year_id,
        "active_school_year_id": active_school_year.id if active_school_year else None,
    }


@router.get("/api/financial-reports/school-years/{school_year_id}")
def get_school_year_detail(
    school_year_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_financial_report_user),
):
    school_year = _ensure_and_reload_school_year(db, school_year_id)
    return _serialize_school_year_detail(db, school_year)


@router.put("/api/financial-reports/reports/{report_id}")
def update_report(
    report_id: int,
    payload: schemas.FinancialReportUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(require_financial_report_user),
):
    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Monthly report not found")
    _validate_school_year_write_allowed(report.school_year)

    updates = payload.model_dump(exclude_unset=True)
    beginning_cash_manual_override = updates.pop("beginning_cash_manual_override", None)
    current_sales_manual_override = updates.pop("current_sales_manual_override", None)

    for field_name, value in updates.items():
        if field_name == "notes":
            setattr(report, field_name, value or "")
        else:
            setattr(report, field_name, _round_money(value))

    if beginning_cash_manual_override is not None:
        report.beginning_cash_manual_override = bool(beginning_cash_manual_override)
    if current_sales_manual_override is not None:
        report.current_sales_manual_override = bool(current_sales_manual_override)

    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_UPDATED",
        details=f"Updated {report.month_name} {report.calendar_year} for {report.school_year.name}",
        request=request,
    )
    db.commit()

    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )

    allocations = sorted(report.school_year.allocations, key=lambda item: (item.sort_order, item.id))
    return {
        "report": _serialize_report(
            report,
            allocations,
            current_sales_auto=_get_report_transaction_sales(db, report),
        )
    }


@router.put("/api/financial-reports/reports/{report_id}/expenses")
def replace_report_expenses(
    report_id: int,
    payload: schemas.FinancialExpensesUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(require_financial_report_user),
):
    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Monthly report not found")
    _validate_school_year_write_allowed(report.school_year)

    for expense in list(report.expenses):
        db.delete(expense)
    db.flush()

    for sort_order, item in enumerate(payload.expenses):
        category = str(item.category or "").strip()
        if not category:
            continue

        db.add(
            models.Expense(
                report_id=report.id,
                category=category,
                amount=_round_money(item.amount),
                sort_order=int(item.sort_order if item.sort_order is not None else sort_order),
            )
        )

    db.flush()
    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )
    _create_default_expenses(db, report)
    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_EXPENSES_UPDATED",
        details=f"Replaced expenses for {report.month_name} {report.calendar_year} in {report.school_year.name}",
        request=request,
    )
    db.commit()

    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )
    allocations = sorted(report.school_year.allocations, key=lambda item: (item.sort_order, item.id))
    return {
        "report": _serialize_report(
            report,
            allocations,
            current_sales_auto=_get_report_transaction_sales(db, report),
        )
    }


@router.put("/api/financial-reports/reports/{report_id}/fund-monitoring")
def replace_report_fund_monitoring(
    report_id: int,
    payload: schemas.FinancialFundMonitoringUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(require_financial_report_user),
):
    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Monthly report not found")
    _validate_school_year_write_allowed(report.school_year)

    for entry in list(report.fund_entries):
        db.delete(entry)
    db.flush()

    valid_category_keys = {
        str(allocation.category_key or "").strip()
        for allocation in report.school_year.allocations
    }

    for item in payload.entries:
        category_key = str(item.category_key or "").strip()
        if not category_key or category_key not in valid_category_keys:
            continue

        db.add(
            models.FundMonitoringEntry(
                report_id=report.id,
                category_key=category_key,
                interest=_round_money(item.interest),
                expenses=_round_money(item.expenses),
                others=_round_money(item.others),
                cash_on_bank=_round_money(item.cash_on_bank),
            )
        )

    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_FUND_MONITORING_UPDATED",
        details=f"Updated fund monitoring for {report.month_name} {report.calendar_year} in {report.school_year.name}",
        request=request,
    )
    db.commit()

    report = (
        db.query(models.MonthlyReport)
        .options(
            joinedload(models.MonthlyReport.expenses),
            joinedload(models.MonthlyReport.fund_entries),
            joinedload(models.MonthlyReport.school_year).joinedload(models.SchoolYear.allocations),
        )
        .filter(models.MonthlyReport.id == report_id)
        .first()
    )
    allocations = sorted(report.school_year.allocations, key=lambda item: (item.sort_order, item.id))
    return {
        "report": _serialize_report(
            report,
            allocations,
            current_sales_auto=_get_report_transaction_sales(db, report),
        )
    }


@router.put("/api/financial-reports/school-years/{school_year_id}/allocations")
def replace_allocations(
    school_year_id: int,
    payload: schemas.FinancialAllocationsUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    school_year = _ensure_and_reload_school_year(db, school_year_id)
    _validate_school_year_write_allowed(school_year)

    for allocation in list(school_year.allocations):
        db.delete(allocation)
    db.flush()

    for sort_order, item in enumerate(payload.allocations):
        label = str(item.label or "").strip()
        category_key = str(item.category_key or "").strip()
        if not label or not category_key:
            continue
        db.add(
            models.Allocation(
                school_year_id=school_year.id,
                category_key=category_key,
                label=label,
                percentage=round(float(item.percentage or 0.0), 2),
                opening_balance=_round_money(item.opening_balance),
                sort_order=int(item.sort_order if item.sort_order is not None else sort_order),
            )
        )

    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_ALLOCATIONS_UPDATED",
        details=f"Updated allocations for {school_year.name}",
        request=request,
    )
    db.commit()

    school_year = _ensure_and_reload_school_year(db, school_year_id)
    detail = _serialize_school_year_detail(db, school_year)
    return {
        "allocations": detail["allocations"],
        "allocation_percent_total": detail["dashboard"]["allocation_percent_total"],
    }


@router.get("/api/financial-reports/school-years/{school_year_id}/export")
def export_school_year_workbook(
    school_year_id: int,
    report_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_financial_report_user),
):
    school_year = _ensure_and_reload_school_year(db, school_year_id)
    export_path = _build_school_year_workbook_export(
        db,
        school_year,
        selected_report_id=report_id,
    )

    return FileResponse(
        export_path,
        filename=f"CANTEEN-REPORT-{school_year.name}.xlsx",
        media_type=EXCEL_MEDIA_TYPE,
        background=BackgroundTask(_remove_file_if_exists, export_path),
    )


@router.get("/api/financial-reports/template")
def download_report_template(
    _: models.User = Depends(require_financial_report_user),
):
    template_path = _template_path()
    if not os.path.isfile(template_path):
        raise HTTPException(status_code=404, detail="Report template file not found")

    return FileResponse(
        template_path,
        filename=os.path.basename(template_path),
        media_type=EXCEL_MEDIA_TYPE,
    )


@router.get("/api/financial-reports/templates/list")
def list_report_templates(
    _: models.User = Depends(require_financial_report_user),
):
    try:
        config = _get_templates_config()
        active_filename = config.get("active_filename", DEFAULT_TEMPLATE_FILENAME)

        registered = {}
        for t in config.get("templates", []):
            if isinstance(t, dict) and "filename" in t:
                registered[t["filename"]] = t

        items = []
        if os.path.isdir(TEMPLATE_DIR):
            for f in sorted(os.listdir(TEMPLATE_DIR)):
                if f.endswith(".xlsx") and not f.startswith("~") and not f.startswith("."):
                    file_path = os.path.join(TEMPLATE_DIR, f)
                    reg_info = registered.get(f, {})

                    try:
                        mtime = os.path.getmtime(file_path)
                        uploaded_at = datetime.fromtimestamp(mtime).isoformat()
                        file_size = os.path.getsize(file_path)
                    except Exception:
                        uploaded_at = "2026-01-01T00:00:00Z"
                        file_size = 0

                    item = {
                        "filename": f,
                        "name": reg_info.get("name") or f.replace(".xlsx", ""),
                        "is_default": f == DEFAULT_TEMPLATE_FILENAME or reg_info.get("is_default", False),
                        "uploaded_at": reg_info.get("uploaded_at") or uploaded_at,
                        "file_size_bytes": file_size,
                        "is_active": f == active_filename,
                    }
                    items.append(item)

        if not items:
            items.append({
                "filename": DEFAULT_TEMPLATE_FILENAME,
                "name": "Default DepEd Canteen Report Template",
                "is_default": True,
                "uploaded_at": "2026-01-01T00:00:00Z",
                "file_size_bytes": 167793,
                "is_active": True,
            })

        return {
            "active_filename": active_filename,
            "templates": items,
        }
    except Exception as exc:
        print(f"Error listing report templates: {exc}")
        return {
            "active_filename": DEFAULT_TEMPLATE_FILENAME,
            "templates": [
                {
                    "filename": DEFAULT_TEMPLATE_FILENAME,
                    "name": "Default DepEd Canteen Report Template",
                    "is_default": True,
                    "uploaded_at": "2026-01-01T00:00:00Z",
                    "file_size_bytes": 167793,
                    "is_active": True,
                }
            ],
        }


@router.post("/api/financial-reports/templates/upload")
def upload_report_template(
    file: UploadFile = File(...),
    set_active: bool = Form(False),
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx) files are allowed")

    clean_filename = os.path.basename(file.filename)
    dest_path = os.path.join(TEMPLATE_DIR, clean_filename)

    try:
        content = file.file.read()
        with ZipFile(BytesIO(content)) as zip_test:
            if "[Content_Types].xml" not in zip_test.namelist():
                raise ValueError("Invalid Excel workbook format")

        with open(dest_path, "wb") as f:
            f.write(content)

        config = _get_templates_config()
        existing = [t for t in config.get("templates", []) if t["filename"] != clean_filename]
        existing.append(
            {
                "filename": clean_filename,
                "name": clean_filename.replace(".xlsx", ""),
                "is_default": clean_filename == DEFAULT_TEMPLATE_FILENAME,
                "uploaded_at": datetime.utcnow().isoformat(),
            }
        )

        if set_active or len(existing) == 1:
            config["active_filename"] = clean_filename
        config["templates"] = existing
        _save_templates_config(config)

        return {
            "message": f"Template '{clean_filename}' uploaded successfully",
            "active_filename": config["active_filename"],
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to process uploaded Excel template: {str(exc)}")


@router.post("/api/financial-reports/templates/select")
def select_report_template(
    payload: dict,
    current: models.User = Depends(auth.require_admin),
):
    target_filename = payload.get("filename")
    if not target_filename:
        raise HTTPException(status_code=400, detail="Filename is required")

    dest_path = os.path.join(TEMPLATE_DIR, target_filename)
    if not os.path.isfile(dest_path):
        raise HTTPException(status_code=404, detail=f"Template file '{target_filename}' not found")

    config = _get_templates_config()
    config["active_filename"] = target_filename
    _save_templates_config(config)

    return {
        "message": f"Active report template updated to '{target_filename}'",
        "active_filename": target_filename,
    }


@router.delete("/api/financial-reports/templates/{filename}")
def delete_report_template(
    filename: str,
    current: models.User = Depends(auth.require_admin),
):
    if filename == DEFAULT_TEMPLATE_FILENAME:
        raise HTTPException(status_code=400, detail="Cannot delete default report template")

    config = _get_templates_config()
    if filename == config.get("active_filename"):
        raise HTTPException(
            status_code=400,
            detail="Cannot delete currently active report template. Please select another active template first.",
        )

    target_path = os.path.join(TEMPLATE_DIR, filename)
    if os.path.isfile(target_path):
        os.remove(target_path)

    config["templates"] = [t for t in config.get("templates", []) if t["filename"] != filename]
    _save_templates_config(config)

    return {"message": f"Template '{filename}' deleted successfully"}


@router.post("/api/financial-reports/backup")
def backup_database(
    request: Request,
    db: Session = Depends(get_db),
    current: models.User = Depends(auth.require_admin),
):
    database_path = _sqlite_database_path()
    if not database_path or not os.path.isfile(database_path):
        raise HTTPException(status_code=400, detail="Database backup is available only for local SQLite deployments")

    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    backup_path = f"{database_path}.backup-{timestamp}-financial-reports"
    shutil.copy2(database_path, backup_path)
    _audit_log(
        db,
        user_id=current.id,
        action="FINANCIAL_REPORT_DATABASE_BACKUP",
        details=os.path.basename(backup_path),
        request=request,
    )
    db.commit()

    return {
        "message": "Database backup created",
        "filename": os.path.basename(backup_path),
        "path": backup_path,
    }
